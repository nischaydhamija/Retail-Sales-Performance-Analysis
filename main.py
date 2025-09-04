"""
Retail Sales Performance Analysis

Analyzes Superstore.csv using SQL queries, pivot tables, and visualizations.
Exports results and dashboard to Excel.
"""
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import pandasql as psql
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import os

# Load dataset
DATA_PATH = 'Superstore.csv'
df = pd.read_csv(DATA_PATH, encoding='utf-8')

# --- SQL Queries ---
# Top 10 products by sales
q1 = """
SELECT [Product Name], SUM([Sales]) AS Total_Sales
FROM df
GROUP BY [Product Name]
ORDER BY Total_Sales DESC
LIMIT 10;
"""
top_products = psql.sqldf(q1, locals())

# Monthly sales trend
q2 = """
SELECT strftime('%Y-%m', [Order Date]) AS Month, SUM([Sales]) AS Monthly_Sales
FROM df
GROUP BY Month
ORDER BY Month;
"""
df['Order Date'] = pd.to_datetime(df['Order Date'])
df['Order Date_str'] = df['Order Date'].dt.strftime('%Y-%m')
monthly_sales = df.groupby('Order Date_str')['Sales'].sum().reset_index()

# Profit by Region and Category
q3 = """
SELECT [Region], [Category], SUM([Profit]) AS Total_Profit
FROM df
GROUP BY [Region], [Category]
ORDER BY [Region], [Category];
"""
profit_region_cat = psql.sqldf(q3, locals())

# --- Pivot Tables ---
pivot_sales = pd.pivot_table(df, index='Region', columns='Category', values='Sales', aggfunc='sum')
pivot_profit = pd.pivot_table(df, index='Region', columns='Category', values='Profit', aggfunc='sum')

# --- Visualizations ---
# Bar chart: Top products
plt.figure(figsize=(10,6))
sns.barplot(x='Total_Sales', y='Product Name', data=top_products, palette='viridis')
plt.title('Top 10 Products by Sales')
plt.tight_layout()
plt.savefig('top_products.png')
plt.close()

# Line chart: Monthly sales trend
plt.figure(figsize=(10,6))
sns.lineplot(x='Order Date_str', y='Sales', data=monthly_sales, marker='o')
plt.title('Monthly Sales Trend')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('monthly_sales.png')
plt.close()

# Pie chart: Profit by Region
region_profit = df.groupby('Region')['Profit'].sum().reset_index()
plt.figure(figsize=(8,8))
plt.pie(region_profit['Profit'], labels=region_profit['Region'], autopct='%1.1f%%', startangle=140)
plt.title('Profit Distribution by Region')
plt.tight_layout()
plt.savefig('profit_region.png')
plt.close()

# --- Export to Excel ---
wb = Workbook()
ws1 = wb.active
ws1.title = 'Top Products'
for r in dataframe_to_rows(top_products, index=False, header=True):
    ws1.append(r)
img1 = XLImage('top_products.png')
ws1.add_image(img1, 'G2')

ws2 = wb.create_sheet('Monthly Sales')
for r in dataframe_to_rows(monthly_sales, index=False, header=True):
    ws2.append(r)
img2 = XLImage('monthly_sales.png')
ws2.add_image(img2, 'D2')

ws3 = wb.create_sheet('Profit by Region/Category')
for r in dataframe_to_rows(profit_region_cat, index=False, header=True):
    ws3.append(r)
img3 = XLImage('profit_region.png')
ws3.add_image(img3, 'E2')

ws4 = wb.create_sheet('Pivot Sales')
for r in dataframe_to_rows(pivot_sales, index=True, header=True):
    ws4.append(r)
ws5 = wb.create_sheet('Pivot Profit')
for r in dataframe_to_rows(pivot_profit, index=True, header=True):
    ws5.append(r)

# --- Summary Sheet ---
ws_summary = wb.create_sheet('Summary')
summary = [
    ['Key Insights'],
    ['- Top selling products are listed in Top Products sheet.'],
    ['- Monthly sales trend shows seasonality and growth.'],
    ['- Profit varies significantly by region and category.'],
    ['- See Pivot tables for detailed breakdowns.'],
    ['- Visualizations embedded in relevant sheets.']
]
for row in summary:
    ws_summary.append(row)

# Save Excel file
os.makedirs('output', exist_ok=True)
wb.save('output/dashboard.xlsx')

print('Analysis complete. Results exported to output/dashboard.xlsx')
